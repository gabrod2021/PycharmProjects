import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.styles import Font
import string

#TABLA PIVOTE CON PANDAS
archivo_excel=pd.read_excel('supermarket_sales.xlsx')
#print(archivo_excel[['Gender', 'Product line', 'Total']])

tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)

tabla_pivote.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')

########################################################################################################################

#AUTOMATIZACION CON PYTHON USANDO OPENPYXL
wb = load_workbook('sales_2021.xlsx')
pestaña = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

"""print(min_col)
print(max_col)
print(min_fila)
print(max_fila)"""

#GRAFICO
barchar = BarChart()

data = Reference(pestaña, min_col=min_col + 1, max_col=max_col, min_row=min_fila, max_row=max_fila)
categorias = Reference(pestaña, min_col=min_col , max_col=min_col, min_row=min_fila + 1, max_row=max_fila)

barchar.add_data(data, titles_from_data=True)
barchar.set_categories(categorias)

pestaña.add_chart(barchar, 'B12')
barchar.title = 'VENTAS'
barchar.style = 2
#pestaña['B8'] = '=SUM(B6:B7)'
#pestaña['B8'].style = 'Currency'

abecedario = list(string.ascii_uppercase)
abecedario_excel = abecedario[0:max_col]

#print(abecedario_excel)


for i in abecedario_excel:
    if i != 'A':
        pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
        pestaña[f'{i}{max_fila+1}'].style = 'Currency'
    #else:
        #pestaña[f'{i}{max_fila + 1}'] = 'TOTAL'
pestaña[f'{abecedario_excel[0]}{max_fila + 1}'] = 'TOTALES:'

pestaña['A1'] = 'REPORTE'
pestaña['A2'] = '2021'

pestaña['A1'].font = Font('Arial', bold=True, size=18)
pestaña['A2'].font = Font('Arial', bold=True, size=15)

wb.save('sales_2021.xlsx')



